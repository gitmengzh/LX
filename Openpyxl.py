#coding:utf-8



from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter


wb = load_workbook(filename = r'C:\Users\meng4\Desktop\test\test.xlsx')
ws = wb.get_sheet_by_name("Sheet1")
wb1 = Workbook()
ewb1 = ExcelWriter(workbook=wb1)
dest_filename = r"C:\Users\meng4\Desktop\test\test1.xlsx"
ws1 = wb1.worksheets[0]
ws1.title = "test"

for i in range(1,36):
    li = []
    for row_num in xrange (1,213824):
        c2 = ws.cell(row = row_num,column = 2).value
        c3 = ws.cell(row = row_num,column = 3).value

        if c2 == i:
            if c3 in li:
                continue
            else:
                ls.append(c3)
        else:
            continue

    ws1.cell(row = i,column = 1).value = i
    ws1.cell(row = i,column = 2).value = i
ewb1.save(filename = dest_filename)

'''

for i in range(0,36):#此处遍历列1寻找0-35
    li=[]#给每个数建立一个list表
    for row_num in xrange(0,213824):#表示遍历的行数
        c2=ws.cell(row=row_num,column=2).value#获取列2对应于excel中的C列的数据
        c3=ws.cell(row=row_num,column=3).value#获取列3对应于excel中的D列的数据
        #对于每个i遍历列C,找出C列跟i相同的数据，并将对应的D列的不同数据保存到li中
        if c2==i:
            if c3 in li:
                continue
            else:
                li.append(c3)
        else:
            continue
#print len(li)
    ws1.cell(row=i,column=0).value=i
    ws1.cell(row=i,column=1).value=len(li)
ewb1.save(filename=dest_filename)#保存一定要有，否则不会有结果



'''